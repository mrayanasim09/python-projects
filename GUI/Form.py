# This code is made by MRayan Asim
# to install these packages:
# pip install openpyxl
import openpyxl
from tkinter import *
from tkinter.messagebox import *

# Declare global variables
name_field = None
course_field = None
sem_field = None
form_no_field = None
contact_no_field = None
email_id_field = None
address_field = None


def excel(file_path):
    # Create or load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    # Create or select the active sheet
    sheet = wb.active
    if sheet.title != "Form Data":
        sheet = wb.create_sheet("Form Data")

    # Resize the width of columns in the spreadsheet
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 10
    sheet.column_dimensions["C"].width = 10
    sheet.column_dimensions["D"].width = 20
    sheet.column_dimensions["E"].width = 20
    sheet.column_dimensions["F"].width = 40
    sheet.column_dimensions["G"].width = 50

    # Write given data to the spreadsheet at a particular location
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Course"
    sheet.cell(row=1, column=3).value = "Semester"
    sheet.cell(row=1, column=4).value = "Form Number"
    sheet.cell(row=1, column=5).value = "Contact Number"
    sheet.cell(row=1, column=6).value = "Email id"
    sheet.cell(row=1, column=7).value = "Address"

    # Save the workbook
    wb.save(file_path)


def focus1(event):
    # set focus on the course_field box
    course_field.focus_set()


def focus2(event):
    # set focus on the sem_field box
    sem_field.focus_set()


def focus3(event):
    # set focus on the form_no_field box
    form_no_field.focus_set()


def focus4(event):
    # set focus on the contact_no_field box
    contact_no_field.focus_set()


def focus5(event):
    # set focus on the email_id_field box
    email_id_field.focus_set()


def focus6(event):
    # set focus on the address_field box
    address_field.focus_set()


def clear():
    # clear the content of text entry boxes
    name_field.delete(0, END)
    course_field.delete(0, END)
    sem_field.delete(0, END)
    form_no_field.delete(0, END)
    contact_no_field.delete(0, END)
    email_id_field.delete(0, END)
    address_field.delete(0, END)


def insert(file_path):
    # if user does not fill any entry, show a message box with an error
    if (
        name_field.get() == ""
        and course_field.get() == ""
        and sem_field.get() == ""
        and form_no_field.get() == ""
        and contact_no_field.get() == ""
        and email_id_field.get() == ""
        and address_field.get() == ""
    ):
        showerror("Error", "Please fill in the form fields.")
    else:
        # Create or load the workbook
        try:
            wb = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            wb = openpyxl.Workbook()

        # Create or select the active sheet
        sheet = wb.active
        if sheet.title != "Form Data":
            sheet = wb.create_sheet("Form Data")

        # Assigning the max row value up to which data is written in the spreadsheet to the variable
        current_row = sheet.max_row

        # Write the form data to the spreadsheet
        sheet.cell(row=current_row + 1, column=1).value = name_field.get()
        sheet.cell(row=current_row + 1, column=2).value = course_field.get()
        sheet.cell(row=current_row + 1, column=3).value = sem_field.get()
        sheet.cell(row=current_row + 1, column=4).value = form_no_field.get()
        sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get()
        sheet.cell(row=current_row + 1, column=6).value = email_id_field.get()
        sheet.cell(row=current_row + 1, column=7).value = address_field.get()

        # Save the workbook
        wb.save(file_path)

        # Set focus on the name_field box
        name_field.focus_set()

        # Call the clear() function
        clear()


if __name__ == "__main__":
    # Create a GUI window
    root = Tk()

    # Set the background color of GUI window
    root.configure(background="light green")

    # Set the title of GUI window
    root.title("Registration Form")

    def submit_form():
        file_path = file_path_entry.get()
        if file_path:
            excel(file_path)
            insert(file_path)
        else:
            showerror("Error", "Please enter a file path.")

    # Create a label for the file path entry
    file_path_label = Label(root, text="File Path", bg="light green")
    file_path_label.grid(row=0, column=0)

    # Create an entry for the file path
    file_path_entry = Entry(root)
    file_path_entry.grid(row=0, column=1, ipadx="100")

    # Create a Submit Button and place it into the root window
    submit = Button(root, text="Submit", fg="Black", bg="Red", command=submit_form)
    submit.grid(row=1, column=1)

    # Start the GUI
    root.mainloop()
